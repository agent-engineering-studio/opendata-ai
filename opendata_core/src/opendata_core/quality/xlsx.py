"""Convertitore XLSX → CSV server-side (#157, follow-up di #101).

Il Quality Lab UI converte gli XLSX client-side (SheetJS); questo motore porta la
STESSA capability lato server per i consumatori REST/A2A. Deterministico: legge il
foglio richiesto (o il primo) e lo serializza in CSV (virgola, quoting minimo),
senza inventare tipi — i valori sono resi come testo così come li vede openpyxl.

Dipendenza opzionale `openpyxl`, importata pigramente (extra `converters`, come
`pyarrow` per Parquet). Senza la libreria la funzione NON solleva: ritorna
`ok=False` con un messaggio chiaro (l'endpoint mappa su 501). Gli `.xls` legacy
(BIFF) non sono supportati da openpyxl → messaggio esplicito, niente crash.
"""

from __future__ import annotations

import csv
import io
from typing import Any

#: Cap difensivo sulle celle serializzate (anti-esplosione memoria su fogli enormi).
_MAX_CELLS = 5_000_000


def _empty(error: str | None = None) -> dict[str, Any]:
    return {
        "ok": False, "content": None, "sheet": None, "sheets": [],
        "righe": 0, "colonne": 0, "warnings": [], "error": error,
    }


def _cell_to_str(value: Any) -> str:
    """Serializza una cella in stringa senza reinterpretarne il tipo.

    openpyxl restituisce già int/float/datetime/bool/str: li rendiamo in modo
    stabile (le date in ISO) e lasciamo il resto invariato. `None` → stringa vuota.
    """
    if value is None:
        return ""
    # datetime/date hanno isoformat(); numeri e bool via str().
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        try:
            return isoformat()
        except Exception:  # noqa: BLE001 — fallback su str, mai bloccante
            return str(value)
    return str(value)


def xlsx_to_csv(data: bytes, *, sheet: str | None = None) -> dict[str, Any]:
    """Converte un workbook XLSX (bytes) in CSV.

    Ritorna `content` (testo CSV), il nome del foglio usato, l'elenco dei fogli,
    righe/colonne e `warnings`. Con `ok=False` il campo `error` spiega il motivo
    (openpyxl assente, file non valido, xls legacy, foglio inesistente, …).
    """
    if not data:
        return _empty("Il file è vuoto.")

    try:
        import openpyxl
    except ImportError:
        return _empty(
            "Conversione XLSX non disponibile su questo server: manca la dipendenza "
            "opzionale openpyxl (extra 'converters')."
        )

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # noqa: BLE001 — file corrotto o .xls legacy → messaggio chiaro
        return _empty(
            f"XLSX non leggibile: {exc}. Nota: gli .xls legacy (formato BIFF) non "
            "sono supportati — riesporta in .xlsx."
        )

    try:
        sheet_names = list(wb.sheetnames)
        if sheet is not None:
            if sheet not in sheet_names:
                return {**_empty(
                    f"Foglio '{sheet}' inesistente. Fogli disponibili: "
                    f"{', '.join(sheet_names)}."), "sheets": sheet_names}
            ws = wb[sheet]
        else:
            ws = wb.active

        warnings: list[str] = []
        buf = io.StringIO()
        writer = csv.writer(buf)
        n_rows = 0
        n_cols = 0
        cells = 0
        for row in ws.iter_rows(values_only=True):
            values = [_cell_to_str(c) for c in row]
            # rimuove le colonne di coda completamente vuote della riga
            while values and values[-1] == "":
                values.pop()
            n_cols = max(n_cols, len(values))
            cells += len(values)
            if cells > _MAX_CELLS:
                return _empty(
                    f"Foglio troppo grande (> {_MAX_CELLS} celle): converti un "
                    "sottoinsieme o usa un formato colonnare."
                )
            writer.writerow(values)
            n_rows += 1
    finally:
        wb.close()

    if n_rows == 0:
        return {**_empty("Il foglio è vuoto."), "sheet": ws.title, "sheets": sheet_names}

    return {
        "ok": True,
        "error": None,
        "content": buf.getvalue(),
        "sheet": ws.title,
        "sheets": sheet_names,
        "righe": n_rows,
        "colonne": n_cols,
        "warnings": warnings,
    }
