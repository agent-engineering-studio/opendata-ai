"""Test dei convertitori binari server-side (#157): XLSX→CSV e Shapefile→GeoJSON."""

from __future__ import annotations

import csv
import io
import zipfile

import pytest

from opendata_core.quality import shapefile_to_geojson, xlsx_to_csv
from opendata_core.quality import shapefile as shp_mod


def _xlsx_bytes(rows, *, title="Foglio1") -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _shapefile_zip(tmp_path, *, epsg: int | None = 4326, points=None) -> bytes:
    import pyproj
    import shapefile

    points = points or [((12.4924, 41.8902), "Roma")]
    stem = str(tmp_path / "data")
    w = shapefile.Writer(stem, shapeType=shapefile.POINT)
    w.field("name", "C")
    for (x, y), name in points:
        w.point(x, y)
        w.record(name)
    w.close()
    z = io.BytesIO()
    with zipfile.ZipFile(z, "w") as zf:
        for ext in (".shp", ".dbf", ".shx"):
            with open(stem + ext, "rb") as fh:
                zf.writestr("data" + ext, fh.read())
        if epsg is not None:
            zf.writestr("data.prj", pyproj.CRS.from_epsg(epsg).to_wkt())
    return z.getvalue()


# ────────────────────────────── XLSX → CSV ─────────────────────────────────


def test_xlsx_to_csv_happy_path() -> None:
    data = _xlsx_bytes([["nome", "eta"], ["Anna", 30], ["Bruno", 25]])
    out = xlsx_to_csv(data)
    assert out["ok"] is True
    assert out["sheet"] == "Foglio1"
    assert out["righe"] == 3 and out["colonne"] == 2
    parsed = list(csv.reader(io.StringIO(out["content"])))
    assert parsed[0] == ["nome", "eta"]
    assert parsed[1] == ["Anna", "30"]


def test_xlsx_to_csv_named_sheet_and_missing() -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "uno"
    wb.active.append(["x"])
    wb.create_sheet("due").append(["y"])
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    assert xlsx_to_csv(data, sheet="due")["content"].strip() == "y"
    miss = xlsx_to_csv(data, sheet="tre")
    assert miss["ok"] is False and "inesistente" in miss["error"]
    assert set(miss["sheets"]) == {"uno", "due"}


def test_xlsx_to_csv_empty_and_invalid() -> None:
    assert xlsx_to_csv(b"")["ok"] is False
    bad = xlsx_to_csv(b"this is not an xlsx file")
    assert bad["ok"] is False and "non leggibile" in bad["error"]


# ────────────────────────── Shapefile → GeoJSON ────────────────────────────


def test_shapefile_wgs84_no_reprojection(tmp_path) -> None:
    out = shapefile_to_geojson(_shapefile_zip(tmp_path, epsg=4326))
    assert out["ok"] is True
    assert out["feature_count"] == 1
    feat = out["geojson"]["features"][0]
    assert feat["geometry"]["type"] == "Point"
    lon, lat = feat["geometry"]["coordinates"][:2]
    assert lon == pytest.approx(12.4924, abs=1e-3) and lat == pytest.approx(41.8902, abs=1e-3)
    assert feat["properties"]["name"] == "Roma"


def test_shapefile_reprojects_from_3857(tmp_path) -> None:
    # (1113194.9, 5012341.7) in EPSG:3857 ≈ (10.0, 41.0) lon/lat
    out = shapefile_to_geojson(
        _shapefile_zip(tmp_path, epsg=3857, points=[((1113194.9, 5012341.7), "P")])
    )
    assert out["ok"] is True
    lon, lat = out["geojson"]["features"][0]["geometry"]["coordinates"][:2]
    assert lon == pytest.approx(10.0, abs=0.01) and lat == pytest.approx(41.0, abs=0.05)


def test_shapefile_missing_prj_warns(tmp_path) -> None:
    out = shapefile_to_geojson(_shapefile_zip(tmp_path, epsg=None))
    assert out["ok"] is True
    assert any(".prj" in w for w in out["warnings"])


def test_shapefile_invalid_zip_and_no_shp(tmp_path) -> None:
    assert shapefile_to_geojson(b"not a zip")["ok"] is False
    empty = io.BytesIO()
    with zipfile.ZipFile(empty, "w") as zf:
        zf.writestr("readme.txt", "hello")
    out = shapefile_to_geojson(empty.getvalue())
    assert out["ok"] is False and ".shp" in out["error"]


def test_shapefile_zip_bomb_guard(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(shp_mod, "MAX_UNCOMPRESSED", 10)  # 10 byte → qualsiasi shp lo supera
    out = shapefile_to_geojson(_shapefile_zip(tmp_path, epsg=4326))
    assert out["ok"] is False and out["zipbomb"] is True
